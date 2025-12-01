from playwright.sync_api import sync_playwright
import pandas as pd
from openpyxl import load_workbook
import time

# Caminho do arquivo Excel
excel_path = r".input_data/excel_containers/CONTAINERS_MOCK.xlsx"

# Carregar dados com pandas
df = pd.read_excel(excel_path, engine='openpyxl')

# Garantir colunas necessárias
if 'STATUSDEVOLUCAO' not in df.columns:
    df['STATUSDEVOLUCAO'] = ""
if 'LOCALIZACAO' not in df.columns:
    df['LOCALIZACAO'] = ""
df['DEVOLUCAOVAZIO'] = df['DEVOLUCAOVAZIO'].astype(str)

# Abrir workbook para manter formatação
wb = load_workbook(excel_path)
ws = wb.active

# Função para atualizar célula mantendo formatação
def atualizar_celula(linha, coluna, valor):
    ws.cell(row=linha + 2, column=coluna, value=valor)  # +2 porque pandas ignora cabeçalho

# Índices das colunas
col_idx = {
    'DEVOLUCAOVAZIO': df.columns.get_loc('DEVOLUCAOVAZIO') + 1,
    'STATUSDEVOLUCAO': df.columns.get_loc('STATUSDEVOLUCAO') + 1,
    'LOCALIZACAO': df.columns.get_loc('LOCALIZACAO') + 1
}

# Contadores
atualizados = 0
nao_devolvidos = 0
erros = 0

log_file = "resultado_log.txt"
open(log_file, "w").close()

def abrir_site(page, url, tentativas=3):
    for i in range(tentativas):
        try:
            print(f"[INFO] Abrindo site (tentativa {i+1})...")
            page.goto(url, timeout=15000)
            page.wait_for_load_state("networkidle")
            return True
        except Exception as e:
            print(f"[ERRO] Tentativa {i+1} falhou: {e}")
            time.sleep(2)
    return False

def buscar_dados(page, container_id):
    try:
        if not abrir_site(page, "https://www.msc.com/pt/tracking"):
            return None, None, None

        # Fechar pop-up de cookies se aparecer
        try:
            page.click("button:has-text('Aceitar tudo')", timeout=5000)
            print("[INFO] Pop-up de cookies fechado.")
        except:
            print("[INFO] Nenhum pop-up encontrado.")

        print(f"[INFO] Buscando container: {container_id}")
        page.wait_for_selector("#trackingNumber", timeout=10000)
        page.fill("#trackingNumber", container_id)

        # Em vez de clicar no botão (que pode levar ao login), vamos usar Enter
        page.keyboard.press("Enter")
        print("[INFO] Pesquisa enviada com Enter.")

        # Espera dinâmica pelos resultados
        page.wait_for_selector("div.msc-flow-tracking__steps", timeout=15000)

        # Itera pelas linhas completas
        rows = page.locator("div.msc-flow-tracking__steps > div")
        data, status, localizacao = None, None, None

        for i in range(rows.count()):
            cells = rows.nth(i).locator("span.data-value").all_inner_texts()
            if len(cells) >= 3 and "Empty to Shipper" in cells[2]:
                data = cells[0]
                localizacao = cells[1]
                status = cells[2]
                break

        return data, status, localizacao
    except Exception as e:
        print(f"[ERRO] Falha ao buscar dados: {e}")
        return None, None, None

# Execução principal
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()

    try:
        for index, row in df.iterrows():
            # Só processa se a célula estiver realmente vazia ou NaN
            valor = str(row['DEVOLUCAOVAZIO']).strip().lower()
            if valor in ["", "nan", "none"]:
                data, status, localizacao = buscar_dados(page, row['CONTAINER'])
                print(f"[RESULTADO] {row['CONTAINER']}: Data={data}, Status={status}, Localização={localizacao}")

                if status and "Empty to Shipper" in status and data:
                    atualizar_celula(index, col_idx['DEVOLUCAOVAZIO'], data)
                    atualizar_celula(index, col_idx['STATUSDEVOLUCAO'], status)
                    atualizar_celula(index, col_idx['LOCALIZACAO'], localizacao if localizacao else "Não encontrado")
                    atualizados += 1
                else:
                    atualizar_celula(index, col_idx['DEVOLUCAOVAZIO'], "Não encontrado")
                    atualizar_celula(index, col_idx['STATUSDEVOLUCAO'], status if status else "Não encontrado")
                    atualizar_celula(index, col_idx['LOCALIZACAO'], localizacao if localizacao else "Não encontrado")
                    nao_devolvidos += 1

                # Salvar imediatamente após cada linha (garantia)
                wb.save(excel_path)

                # Log
                with open(log_file, "a", encoding="utf-8") as f:
                    f.write(f"{row['CONTAINER']}: Data={data}, Status={status}, Localização={localizacao}\n")

    except Exception as e:
        print(f"[ERRO GERAL] {e}")
    finally:
        wb.save(excel_path)
        browser.close()

print("\nResumo:")
print(f"Atualizados: {atualizados}")
print(f"Ainda não devolvidos: {nao_devolvidos}")
print(f"Erros: {erros}")
print(f"Log salvo em: {log_file}")
print("Processo concluído!")