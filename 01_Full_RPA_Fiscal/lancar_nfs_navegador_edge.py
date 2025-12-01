
from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, WebDriverException, NoSuchElementException
import pandas as pd
from openpyxl import load_workbook
import time

# --- Configurações Iniciais ---
url_rede = "https://sistema-fiscal-ficticio.com/login"
caminho_excel = r"./input_data/excel_protocolo/PROTOCOLO_MOCK.xlsx"

from selenium.webdriver.edge.service import Service

def iniciar_driver():
    """Inicializa o driver do Microsoft Edge com caminho direto."""
    try:
        print("Abrindo o sistema com Microsoft Edge...")
        edge_options = EdgeOptions()
        edge_options.use_chromium = True

        # Caminho direto para o msedgedriver.exe
        caminho_driver = r"C:\Users\camila\OneDrive - empresa\Desktop\Codigos\msedgedriver.exe"
        servico = Service(caminho_driver)

        driver = webdriver.Edge(service=servico, options=edge_options)
        driver.maximize_window()
        return driver
    except WebDriverException:
        print("Erro: Não foi possível iniciar o driver do Microsoft Edge. Verifique se o caminho está correto e se o driver é compatível com sua versão do navegador.")
        return None

def ler_planilha(caminho_excel):
    """Lê os dados da planilha e retorna o DataFrame e o Workbook."""
    try:
        print("Lendo dados da planilha...")
        df = pd.read_excel(caminho_excel, header=None, engine='openpyxl', dtype={8: str, 9: str})
        wb = load_workbook(caminho_excel)
        return df, wb
    except FileNotFoundError:
        print(f"Erro: O arquivo não foi encontrado em {caminho_excel}.")
        return None, None
    except Exception as e:
        print(f"Erro ao ler a planilha: {e}")
        return None, None

def tentar_acao(driver, by, element, action, value=None, retries=3):
    """Executa uma ação (clique ou preenchimento) com tratamento de erro e saída mínima."""
    for attempt in range(retries):
        try:
            wait = WebDriverWait(driver, 10)
            elem = wait.until(EC.element_to_be_clickable((by, element)))
            if action == 'click':
                elem.click()
            elif action == 'send_keys':
                elem.clear()
                elem.send_keys(value)
            return True
        except (TimeoutException, StaleElementReferenceException):
            if attempt == retries - 1:
                print(f"Erro: Falha ao encontrar ou interagir com o elemento '{element}'.")
            time.sleep(2)
        except Exception as e:
            print(f"Erro inesperado ao realizar ação em '{element}': {e}")
            return False
    return False

# --- Fluxo Principal ---
if __name__ == "__main__":
    driver = iniciar_driver()
    if not driver:
        exit()

    df, wb = ler_planilha(caminho_excel)
    if df is None:
        driver.quit()
        exit()

    ws = wb.active
    
    driver.get(url_rede)
    time.sleep(5)

    # Ler a planilha e processar as notas
    for i in range(2, len(df)):
        try:
            if ws.cell(row=i + 1, column=11).value is not None:
                print(f"Linha {i + 1} já tem ID de lançamento, pulando...")
                continue
            if pd.isnull(df.iloc[i, 0]):
                print(f"Linha {i + 1} sem data de emissão, pulando...")
                continue

            data_emissao = pd.to_datetime(df.iloc[i, 0]).strftime('%d/%m/%Y')
            numero_nota = str(df.iloc[i, 2])
            print(f"\n--- Processando nota {numero_nota} da linha {i + 1} ---")
            
            # 01. Clicar em nota fiscal
            if not tentar_acao(driver, By.XPATH, "/html/body/form/div[5]/div/ul/li[2]/a", 'click'):
                continue
            time.sleep(2)

            # 02. Clicar em incluir nota fiscal
            if not tentar_acao(driver, By.XPATH, "/html/body/form/div[5]/div/ul/li[2]/ul/li[1]/ul/li[1]/a", 'click'):
                continue
            time.sleep(2)

            # 03. Clicar em pesquisar requisitante
            if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[3]/input[2]", 'click'):
                continue
            time.sleep(2)

            # 04-08. Preencher e selecionar CWS
            try:
                Select(WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[2]/select")))).select_by_index(2)
                driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[4]/input").send_keys("camila")
                driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[5]/input").click()
                time.sleep(2)
                driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[2]/div/div/fieldset[2]/table/tbody/tr/td/div/table/tbody/tr[2]/td[1]/input").click()
                time.sleep(1)
                driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[2]/center/input[1]").click()
                time.sleep(2)
            except Exception as e:
                print(f"Erro ao preencher dados do requisitante: {e}")
                continue

            # 09-12. Selecionar e preencher dados da nota
            try:
                Select(driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/select[1]")).select_by_index(3)
                Select(driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/select[5]")).select_by_index(4)
                
                valor_total_raw = df.iloc[i, 4]
                valor_total = f"{valor_total_raw:.2f}".replace(".", ",")
                
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/input[2]", 'send_keys', numero_nota):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/input[5]", 'send_keys', valor_total):
                    continue

            except Exception as e:
                print(f"Erro ao preencher dados da nota: {e}")
                continue

            # 13. Preencher data de emissão
            xpath_emissao = "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/input[7]"
            campo_emissao = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath_emissao)))
            campo_emissao.send_keys(Keys.CONTROL + "a")
            campo_emissao.send_keys(Keys.BACKSPACE)
            campo_emissao.send_keys(data_emissao)

            # 14. Preencher data de recebimento
            data_recebimento_raw = df.iloc[i, 5]
            data_recebimento = pd.to_datetime(data_recebimento_raw).strftime('%d/%m/%Y') if pd.notnull(data_recebimento_raw) else ""
            xpath_recebimento = "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/input[8]"
            campo_recebimento = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath_recebimento)))
            campo_recebimento.send_keys(Keys.CONTROL + "a")
            campo_recebimento.send_keys(Keys.BACKSPACE)
            campo_recebimento.send_keys(data_recebimento)
            time.sleep(1)

            # 15-16. Preencher Contrato e Observação
            contrato = str(df.iloc[1, 12])
            observacao = str(df.iloc[i, 6])
            if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/div/input", 'send_keys', contrato):
                continue
            if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/center/textarea", 'send_keys', observacao):
                continue

            # 17-22. Lógica para Fornecedor
            if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[6]/input[2]", 'click'):
                continue
            time.sleep(2)
            
            fornecedor_codigo = str(df.iloc[1, 11])
            try:
                Select(WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[1]/div/div/fieldset[1]/table[1]/tbody/tr/td[3]/select")))).select_by_index(2)
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[1]/div/div/fieldset[1]/table[1]/tbody/tr/td[5]/input", 'send_keys', fornecedor_codigo):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[1]/div/div/fieldset[1]/table[1]/tbody/tr/td[6]/input", 'click'):
                    continue
                time.sleep(2)
                
                linhas_resultado = driver.find_elements(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[1]/div/div/fieldset[2]/table/tbody/tr/td/div/table/tbody/tr")
                if len(linhas_resultado) >= 2:
                    if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[1]/div/div/fieldset[2]/table/tbody/tr/td/div/table/tbody/tr[2]/td[1]/input", 'click'):
                        continue
                    time.sleep(1)
                    if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/div[1]/center/input[1]", 'click'):
                        continue
                    time.sleep(2)
                else:
                    print("Aviso: Nenhum resultado encontrado para o fornecedor. Pulando para a próxima nota.")
                    continue
            except Exception as e:
                print(f"Erro ao processar o fornecedor: {e}")
                continue

            # 23-33. Preenchimento de dados de custo e adição
            try:
                Select(driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/select[1]")).select_by_index(3)
                time.sleep(2)
                Select(driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/select[2]")).select_by_index(182)
                time.sleep(2)

                conta = str(df.iloc[i, 7])
                sub1 = df.iloc[i, 8].zfill(3)
                sub2 = df.iloc[i, 9].zfill(2)

                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/fieldset[1]/input[1]", 'send_keys', observacao):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/fieldset[2]/input[1]", 'send_keys', conta):
                    continue
                
                # PAUSA CRÍTICA ADICIONADA AQUI
                time.sleep(1)

                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/fieldset[2]/input[2]", 'send_keys', sub1):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/fieldset[2]/input[3]", 'send_keys', sub2):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/fieldset[2]/input[4]", 'send_keys', ""):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/fieldset[3]/input[1]", 'send_keys', "1"):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/fieldset[3]/input[2]", 'send_keys', "1"):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/fieldset[3]/input[3]", 'send_keys', valor_total):
                    continue
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[7]/div/center/input", 'click'):
                    continue
                time.sleep(2)
            except Exception as e:
                print(f"Erro ao preencher dados de custo: {e}")
                continue

            # 34-36. Anexar documento
            try:
                Select(driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/div/select")).select_by_index(4)
                campo_arquivo = driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/div/input")
                caminho_arquivo_completo = rf"C:\Users\camila\Desktop\Códigos Git\01_Full_RPA_Fiscal\{numero_nota}.pdf"
                campo_arquivo.send_keys(caminho_arquivo_completo)
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[4]/div/center/input", 'click'):
                    continue
                time.sleep(2)
            except Exception as e:
                print(f"Erro ao anexar documento: {e}")
                continue

            # 37-39. Finalizar lançamento
            try:
                Select(driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[10]/select")).select_by_index(2)
                time.sleep(1)
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[11]/center/input", 'click'):
                    continue
                time.sleep(1)
                if not tentar_acao(driver, By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[12]/center/input", 'click'):
                    continue
                time.sleep(5)
            except Exception as e:
                print(f"Erro ao finalizar lançamento: {e}")
                continue

            # 40. Clicar em OK no popup de sucesso
            if not tentar_acao(driver, By.XPATH, "/html/body/form/div[8]/div[3]/div/button", 'click'):
                continue
            time.sleep(5)

            # 41-42. Copiar RR e salvar na planilha
            try:
                id_lancamento = driver.find_element(By.XPATH, "/html/body/form/div[7]/div[1]/div[1]/fieldset[5]/input[1]").get_attribute("value")
                ws.cell(row=i + 1, column=11).value = id_lancamento
                wb.save(caminho_excel)
                print(f"Nota {numero_nota} lançada com sucesso. ID: {id_lancamento}")
            except Exception as e:
                print(f"Erro ao capturar o ID de lançamento: {e}")
                continue
            
        except Exception as e:
            print(f"Erro ao processar a nota da linha {i+1} ({numero_nota}): {e}")
            continue

    print("\nProcesso finalizado.")
    driver.quit()
    wb.close()